import openpyxl

class SpreadsheetWriter(object):
    def create_workbook(self, spreadsheet_filename, sheet_name_list):
        self.wb = openpyxl.Workbook()
        self.spreadsheet_filename = spreadsheet_filename

        counter = 0
        self.ws_dict = {}
        for sheet_name in sheet_name_list:
            if counter == 0:
                ws = self.wb.worksheets[0]
            else:
                ws = self.wb.create_sheet()
            counter = counter + 1
            ws.title = sheet_name
            self.ws_dict[ws.title] = ws

    def write_to_sheet(self, ws_name, header_name_list, data, is_use_number_row=True):
            ws = self.ws_dict[ws_name]

            # Writing header
            if is_use_number_row and 'No' not in header_name_list:
                header_name_list = ['No'] + header_name_list
            for i in range(len(header_name_list)):
                ws.cell(row=1, column=i+1).value=header_name_list[i]

            # Writing row data
            row_num = 1
            if is_use_number_row:
                for i in range(len(data)):
                    for j in range(len(data[i])+1):
                        if j == 0:
                            ws.cell(row=i+2, column=j+1).value=row_num
                            row_num += 1
                        else:
                            ws.cell(row=i+2, column=j+1).value=data[i][j-1]
            else:
                for i in range(len(data)):
                    for j in range(len(data[i])):
                        ws.cell(row=i+2, column=j+1).value=data[i][j]
            self.save()

    def convert_data_into_2d_list(self, mode, data):
        two_dim_list = []
        if mode == 'new_user_says':
            for intent_name, rows in data.items():
                for row in rows:
                    two_dim_list.append([row['sheet_name'], intent_name, row['user_says']])
        elif mode == 'classification_report':
            for intent_name, rows in data.items():
                for label, value in rows.items():
                    two_dim_list.append([intent_name, label, value])
        elif mode == 'classification_report_multi_agent':
            for agent_name, datum in data.items():
                for intent_name, rows in datum.items():
                    for label, value in rows.items():
                        two_dim_list.append([agent_name, intent_name, label, value])
        elif mode == 'sync_between_agent':
            for intent_name, user_says_list in data.items():
                for user_says in user_says_list:
                    two_dim_list.append([intent_name, user_says['user_says'].replace('\n', ' ')])
        return two_dim_list


    def save(self):
        self.wb.save(self.spreadsheet_filename)

        